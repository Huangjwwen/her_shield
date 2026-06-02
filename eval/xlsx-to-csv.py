#!/usr/bin/env python
# 把 annotation_filled.xlsx 转成 UTF-8 BOM CSV,便于跑 validate-annotation.js
import openpyxl, csv, sys, io

src = sys.argv[1] if len(sys.argv) > 1 else 'D:/WECHAT/xwechat_files/wxid_s8hy5mq5jblg22_c085/msg/file/2026-06/annotation_filled.xlsx'
dst = sys.argv[2] if len(sys.argv) > 2 else 'eval/annotation_filled_reviewed.csv'

wb = openpyxl.load_workbook(src, data_only=True)
ws = wb.active

with open(dst, 'w', encoding='utf-8-sig', newline='') as f:
    writer = csv.writer(f, lineterminator='\r\n')
    for row in ws.iter_rows(values_only=True):
        writer.writerow(['' if v is None else str(v) for v in row])

print(f'已写入: {dst}')
print(f'  来源: {src}')
print(f'  尺寸: {ws.max_row} 行 × {ws.max_column} 列')
